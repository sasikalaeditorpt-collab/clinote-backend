import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


def build_rich_text_cell(runs):
    """
    Build a rich-text cell from a list of runs.

    Each run is expected to be a dict like:
      {
        "text": "some text",
        "color": "FF0000",   # optional, defaults to black
        "bold": True/False   # optional, defaults to False
      }

    Returns a CellRichText or None.
    """
    if not runs:
        return None

    blocks = []
    for run in runs:
        text = run.get("text", "")

        # Ensure we never pass None into InlineFont
        color = run.get("color") or "000000"
        bold = bool(run.get("bold", False))

        font = InlineFont(color=color, b=bold)
        blocks.append(TextBlock(text=text, font=font))

    return CellRichText(blocks)


def autosize_columns(ws):
    """
    Auto-size columns based on max length of cell values.
    """
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column  # 1-based index
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, CellRichText):
                text_value = "".join(block.text for block in value)
            else:
                text_value = str(value)
            max_length = max(max_length, len(text_value))
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width


def write_excel_summary(all_rows, unmatched):
    """
    Write the audit summary Excel file.

    all_rows: list of dicts, each like:
      {
        "tracking_number": str,
        "patient": str,
        "typist": str,
        "typed": str,
        "dictated": str,
        "T": str,
        "D": str,
        "typed_runs": list (optional),
        "dictated_runs": list (optional),
      }

    unmatched: list of strings (unmatched file paths or names)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Summary"

    # Header row
    headers = [
        "Tracking Number",
        "Patient",
        "Typist",
        "Typed (Rich)",
        "Dictated (Rich)",
        "Typed (Plain)",
        "Dictated (Plain)",
    ]
    ws.append(headers)

    # Center header alignment
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row in all_rows:
        tracking_number = row.get("tracking_number", "")
        patient = row.get("patient", "")
        typist = row.get("typist", "")

        typed_plain = row.get("typed", "")
        dictated_plain = row.get("dictated", "")

        typed_runs = row.get("typed_runs")
        dictated_runs = row.get("dictated_runs")

        # Build rich-text cells if runs are present; otherwise use plain text
        typed_rich = build_rich_text_cell(typed_runs) if typed_runs else typed_plain
        dictated_rich = build_rich_text_cell(dictated_runs) if dictated_runs else dictated_plain

        ws.append([
            tracking_number,
            patient,
            typist,
            typed_rich,
            dictated_rich,
            typed_plain,
            dictated_plain,
        ])

    # Optional: add unmatched sheet
    if unmatched:
        ws_unmatched = wb.create_sheet(title="Unmatched Files")
        ws_unmatched.append(["Unmatched File"])
        for item in unmatched:
            ws_unmatched.append([item])
        autosize_columns(ws_unmatched)

    autosize_columns(ws)

    # Output path
    os.makedirs("output", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join("output", f"audit_summary_{timestamp}.xlsx")

    wb.save(excel_path)
    return excel_path